# coding=UTF-8
import openpyxl
from win32com.client import Dispatch
import os

def write_excel_xlsx(path, sheet_name, value):
    #index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    """
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=value[i][j])
    """
    sheet.cell(6, 1, 'A2')
    sheet.cell(6, 2, 84)
    sheet.cell(6, 1, 76)
    sheet.cell(6, 1, 56)
    sheet.cell(6, 1, 89)
    workbook.save(path)
    print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

def SaveExcel(filename):
    pwd = os.getcwd()
    xlApp = Dispatch('Excel.Application')
    xlApp.Visible = True
    xlBook = xlApp.Workbooks.Open(pwd + '\\' + filename)
    sht = xlBook.Worksheets(1)
    print(sht)
    xlBook.Save()
    xlBook.Close()
    xlApp.Quit()

book_name_xlsx = '公式计算结果验证.xlsx'

sheet_name_xlsx = 'xlsx格式测试表'

value3 = [["姓名", "数学", "语文", "英语", "体育"],
          ["A1", "78", "66", "88", "80"],
          ["B2", "90", "55", "56", "22"],
          ["C3", "96", "75", "89", "67"], ]

write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
SaveExcel(book_name_xlsx)
read_excel_xlsx(book_name_xlsx, sheet_name_xlsx)
