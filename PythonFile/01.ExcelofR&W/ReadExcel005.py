import openpyxl
from win32com.client import Dispatch
import os

def writeToExcel(Excel_file):
    wb1 = openpyxl.load_workbook(Excel_file)
    ws1 = wb1.get_sheet_by_name("xlsx格式测试表")
    print(ws1)
    ws1["A4"].value = "JackLee"
    ws1["B4"].value = 990
    ws1["C4"].value = 578
    ws1["D4"].value = 423
    ws1["E4"].value = 234
    wb1.save(Excel_file)
    print("写入数据成功！")

def SaveExcel(filename):
    pwd = os.getcwd()
    xlApp = Dispatch('Excel.Application')
    xlApp.Visible = True
    xlBook = xlApp.Workbooks.Open(pwd + '\\' + filename)
    sht = xlBook.Worksheets(1)
    print(sht)
    xlBook.Save()
    xlBook.Close()
    xlApp.Quit()

def SaveExcel000():
    xlApp = Dispatch('Excel.Application')

    xlApp.Visible = True

    xlApp.Workbooks.Open(r"C:\Users\lijik\PycharmProjects\PythonFile\01.ExcelofR&W\win32_learning\\公式计算结果验证.xlsx")
    #xlApp.Quit()
    Sheet=None
    Doc=None
    app=None
    while not app==None:
        pythoncom.PumpWaitingMessages()

    xlApp.Save
    print("保存成功！")
    #xlApp.Quit()


def ReadExcel001(Excel_file):
    #just_open(Excel_file)
    wb = openpyxl.load_workbook(Excel_file, data_only=True)
    ws = wb.get_sheet_by_name("xlsx格式测试表")
    a1 = ws.cell(8, 1).value
    a2 = ws.cell(10, 6).value
    print(a1, a2)

if __name__ == "__main__":
    writeToExcel("公式计算结果验证.xlsx")
    #SaveExcel()
    #just_open("公式计算结果验证.xls")
    #ReadExcel("公式计算结果验证.xlsx")