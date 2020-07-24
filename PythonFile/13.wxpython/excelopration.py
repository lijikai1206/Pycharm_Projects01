import openpyxl

def read_excel():
    workbook = openpyxl.load_workbook('data.xlsx')
    worksheet = workbook['Sheet1']
    row3 = [item.value for item in list(worksheet.rows)[2]]
    print('第3行值', row3)
    col3 = [item.value for item in list(worksheet.columns)[2]]
    print('第3行值', col3)
    cell_2_3 = worksheet.cell(row=2, column=3).value
    print('第2行第3列值', cell_2_3)
    max_row = worksheet.max_row
    print('最大行', max_row)

def write():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = 'hi,wwu'
    workbook.save('new.xlsx')

def write_excel(filename='data.xlsx'):
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook['Sheet1']
    worksheet['A1'] = 'hi,wwu'
    workbook.save(filename)


if __name__ == '__main__':
    write_excel(filename='data.xlsx')